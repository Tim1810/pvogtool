import json
from io import BytesIO
from typing import Any

import requests
import streamlit as st
import openpyxl

from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter


# ============================================================
# KONFIGURATION
# ============================================================

PVOG_BASE = "https://pvog.fitko.net/suchdienst/api"

LEIKA_IDS = [
    "99008001014000",
    "99008001014001",
    "99008001014002",
    "99008002010000",
    "99008002010001",
    "99008002010003",
    "99008002010002",
]

# Aktueller produktiver PVOG-Endpunkt
SERVICE_DETAIL_URL = (
    PVOG_BASE +
    "/v6/servicedescriptions/{lbid}/detail"
)

# Aktuelle Beta-API:
# Onlinedienste innerhalb eines ARS, optional nach LeiKa
ONLINE_SERVICES_URL = (
    PVOG_BASE +
    "/v1beta2/onlineservices"
)

# Zuständige OE eines konkreten Onlinedienstes
OD_ORGANISATIONS_URL = (
    PVOG_BASE +
    "/v1beta2/onlineservices/{od_id}/organisations"
)

# OE-Detail
OE_DETAIL_URL = (
    PVOG_BASE +
    "/v5/organisationunits/detail"
)

# OD-Detail
OD_DETAIL_URL = (
    PVOG_BASE +
    "/v2/onlineservices/detail"
)


# ============================================================
# STREAMLIT
# ============================================================

st.set_page_config(
    page_title="PVOG LeiKa-Auswertung",
    page_icon="🏛️",
    layout="wide",
)

st.title("🏛️ PVOG LeiKa-Auswertung")

st.write(
    "Für eine ARS werden die gewünschten PVOG-Daten "
    "für sieben fest definierte LeiKa-IDs ermittelt "
    "und als Excel-Datei ausgegeben."
)


# ============================================================
# HTTP
# ============================================================

SESSION = requests.Session()

SESSION.headers.update({
    "User-Agent": "PVOG-Leika-Streamlit/1.0",
    "Accept": "application/json",
})


def get_json(
    url: str,
    params: dict | None = None,
    timeout: int = 30,
):
    try:
        response = SESSION.get(
            url,
            params=params,
            timeout=timeout,
        )

        if response.status_code != 200:
            return None

        return response.json()

    except Exception:
        return None


# ============================================================
# ALLGEMEINE JSON-HILFSFUNKTIONEN
# ============================================================

def text(value: Any) -> str:

    if value is None:
        return ""

    if isinstance(value, str):
        return value.strip()

    if isinstance(value, (int, float, bool)):
        return str(value)

    return ""


def unique(values):

    result = []

    for value in values:

        value = text(value)

        if value and value not in result:
            result.append(value)

    return result


def walk_json(obj):

    if isinstance(obj, dict):

        yield obj

        for value in obj.values():
            yield from walk_json(value)

    elif isinstance(obj, list):

        for item in obj:
            yield from walk_json(item)


# ============================================================
# ID-ERKENNUNG
# ============================================================

def is_oe_id(value):

    return (
        isinstance(value, str)
        and ".OE." in value
    )


def is_od_id(value):

    return (
        isinstance(value, str)
        and ".OD." in value
    )


def find_oe_ids(obj):

    result = []

    for item in walk_json(obj):

        for key, value in item.items():

            if isinstance(value, str):

                if is_oe_id(value):
                    result.append(value)

    return unique(result)


def find_od_ids(obj):

    result = []

    for item in walk_json(obj):

        for key, value in item.items():

            if isinstance(value, str):

                if is_od_id(value):
                    result.append(value)

    return unique(result)


# ============================================================
# 1. ONLINEDIENSTE FÜR ARS + LEIKA
# ============================================================

def get_online_services(
    ars,
    leika,
):

    return get_json(
        ONLINE_SERVICES_URL,
        params={
            "ars": ars,
            "leikaIds": leika,
        },
    )


# ============================================================
# LBID AUS PVOG-ANTWORT ERMITTELN
# ============================================================

def find_lbid(
    data,
    leika,
):

    candidates = []

    for item in walk_json(data):

        # mögliche LBID-Felder
        for key, value in item.items():

            key_lower = key.lower()

            if (
                "lbid" in key_lower
                and isinstance(value, str)
                and value.strip()
            ):

                candidates.append(
                    value.strip()
                )

        # ID eines Leistungsobjekts
        #
        # Nur aufnehmen, wenn dasselbe Objekt
        # auch die gesuchte LeiKa enthält.

        contains_leika = (
            leika in json.dumps(
                item,
                ensure_ascii=False
            )
        )

        if contains_leika:

            value = item.get("id")

            if (
                isinstance(value, str)
                and value.strip()
                and not is_od_id(value)
                and not is_oe_id(value)
            ):

                candidates.append(
                    value.strip()
                )

    candidates = unique(candidates)

    return candidates[0] if candidates else ""


# ============================================================
# 2. LEISTUNGSDETAIL
# ============================================================

def get_service_detail(
    lbid,
    ars,
):

    if not lbid:
        return None

    url = SERVICE_DETAIL_URL.format(
        lbid=lbid
    )

    return get_json(
        url,
        params={
            "ars": ars
        },
    )


# ============================================================
# 3. OE-DETAIL
# ============================================================

def get_oe_detail(
    oe_id,
    lbid,
    ars,
):

    return get_json(
        OE_DETAIL_URL,
        params={
            "q": oe_id,
            "lbId": lbid,
            "ars": ars,
        },
    )


# ============================================================
# OE-TITEL
# ============================================================

def extract_oe_title(
    data
):

    """
    Nur echte Titel-/Namensfelder eines
    OE-Objektes werden berücksichtigt.

    Es wird NICHT mehr:
        clean_texts[1]
    verwendet.
    """

    if not isinstance(data, dict):
        return ""

    preferred = [
        "title",
        "name",
        "organisationUnitTitle",
        "organisationUnitName",
        "bezeichnung",
    ]

    for key in preferred:

        value = data.get(key)

        if isinstance(value, str):

            value = value.strip()

            if value:
                return value

    # mögliche verschachtelte OE
    for key in [
        "organisationUnit",
        "organisationseinheit",
        "organisation",
    ]:

        child = data.get(key)

        if isinstance(child, dict):

            result = extract_oe_title(
                child
            )

            if result:
                return result

    return ""


# ============================================================
# OE-SIGNATUR-TITEL
# ============================================================

def extract_signature_title(
    data
):

    """
    Sucht ausschließlich nach Feldern,
    die explizit eine Destination-/Signatur-
    Bezeichnung repräsentieren.

    Allgemeine 'name'-Felder werden hier
    NICHT verwendet.
    """

    possible_keys = [
        "destinationSignatureTitle",
        "destinationSignatureName",
        "signatureTitle",
        "signatureName",
        "destinationName",
        "destinationTitle",
        "destinationSignature",
    ]

    for item in walk_json(data):

        for key in possible_keys:

            value = item.get(key)

            if isinstance(value, str):

                value = value.strip()

                if value:
                    return value

            elif isinstance(value, dict):

                for nested_key in [
                    "title",
                    "name",
                    "bezeichnung",
                    "value",
                ]:

                    nested = value.get(
                        nested_key
                    )

                    if isinstance(
                        nested,
                        str
                    ) and nested.strip():

                        return nested.strip()

    return ""


# ============================================================
# 4. OD-ORGANISATIONEN
# ============================================================

def get_od_organisations(
    od_id
):

    url = OD_ORGANISATIONS_URL.format(
        od_id=od_id
    )

    return get_json(
        url
    )


# ============================================================
# OD-DATEN
# ============================================================

def get_od_detail(
    od_id
):

    return get_json(
        OD_DETAIL_URL,
        params={
            "q": od_id
        },
    )


def extract_od_url(
    data
):

    urls = []

    for item in walk_json(data):

        for key, value in item.items():

            if not isinstance(
                value,
                str
            ):
                continue

            value = value.strip()

            if not value.startswith(
                "http"
            ):
                continue

            key_lower = key.lower()

            # Datenschutz separat behandeln
            if (
                "datenschutz" not in key_lower
                and "privacy" not in key_lower
            ):

                urls.append(value)

    return unique(urls)


def extract_datenschutz_url(
    data
):

    urls = []

    for item in walk_json(data):

        for key, value in item.items():

            if not isinstance(
                value,
                str
            ):
                continue

            value = value.strip()

            if not value.startswith(
                "http"
            ):
                continue

            key_lower = key.lower()

            if (
                "datenschutz" in key_lower
                or "privacy" in key_lower
            ):

                urls.append(value)

    return unique(urls)


# ============================================================
# OE-ID-SIGNATUR BEI ROUTING 2.3
# ============================================================

def extract_routing_23_signature(
    data
):

    """
    PVOG/XZuFi 2.3:

    Es werden idSekundaer-Werte aus dem
    XZuFi-/JZuFi-Datensatz gesucht.

    Es wird ausschließlich ein tatsächlich
    vorhandener Wert ausgegeben.
    """

    values = []

    for item in walk_json(data):

        if "idSekundaer" not in item:
            continue

        value = item.get(
            "idSekundaer"
        )

        if isinstance(
            value,
            list
        ):

            for entry in value:

                if isinstance(
                    entry,
                    dict
                ):

                    candidate = entry.get(
                        "value"
                    )

                    if candidate:
                        values.append(
                            str(candidate).strip()
                        )

                elif isinstance(
                    entry,
                    str
                ):

                    values.append(
                        entry.strip()
                    )

        elif isinstance(
            value,
            dict
        ):

            candidate = value.get(
                "value"
            )

            if candidate:
                values.append(
                    str(candidate).strip()
                )

        elif isinstance(
            value,
            str
        ):

            values.append(
                value.strip()
            )

    return unique(values)


# ============================================================
# OE-DATEN EINER LEIKA
# ============================================================

def collect_oe_data(
    ars,
    lbid,
    service_data,
    od_ids,
):

    oe_ids = find_oe_ids(
        service_data
    )

    # Zusätzlich OEs aus den konkreten
    # OD-Zuständigkeitsrelationen laden.
    for od_id in od_ids:

        od_org_data = (
            get_od_organisations(
                od_id
            )
        )

        if od_org_data:

            oe_ids.extend(
                find_oe_ids(
                    od_org_data
                )
            )

    oe_ids = unique(
        oe_ids
    )

    titles = []
    signature_titles = []
    routing_signatures = []

    for oe_id in oe_ids:

        oe_detail = get_oe_detail(
            oe_id,
            lbid,
            ars
        )

        if not oe_detail:
            continue

        # ----------------------------------------------------
        # OE-ID Titel
        # ----------------------------------------------------

        title = extract_oe_title(
            oe_detail
        )

        if title:
            titles.append(title)

        # ----------------------------------------------------
        # OE-ID Signatur Titel
        # ----------------------------------------------------

        signature_title = (
            extract_signature_title(
                oe_detail
            )
        )

        if signature_title:
            signature_titles.append(
                signature_title
            )

        # ----------------------------------------------------
        # OE-ID Signatur bei Routing 2.3
        # ----------------------------------------------------

        routing_values = (
            extract_routing_23_signature(
                oe_detail
            )
        )

        routing_signatures.extend(
            routing_values
        )

    return {
        "oe_ids": unique(oe_ids),
        "oe_titles": unique(titles),
        "signature_titles":
            unique(signature_titles),
        "routing_signatures":
            unique(routing_signatures),
    }


# ============================================================
# OD-DATEN EINER LEIKA
# ============================================================

def collect_od_data(
    od_ids
):

    od_urls = []
    privacy_urls = []

    for od_id in od_ids:

        data = get_od_detail(
            od_id
        )

        if not data:
            continue

        od_urls.extend(
            extract_od_url(
                data
            )
        )

        privacy_urls.extend(
            extract_datenschutz_url(
                data
            )
        )

    return {
        "od_ids":
            unique(od_ids),

        "privacy_urls":
            unique(privacy_urls),

        "od_urls":
            unique(od_urls),
    }


# ============================================================
# EINE LEIKA AUSWERTEN
# ============================================================

def analyse_leika(
    ars,
    leika
):

    # --------------------------------------------------------
    # Schritt 1:
    # Onlinedienste für ARS + LeiKa
    # --------------------------------------------------------

    online_data = (
        get_online_services(
            ars,
            leika
        )
    )

    if not online_data:

        return {
            "Leika-ID": leika,
            "OE-ID": "",
            "OE-ID Titel": "",
            "OE-ID Signatur Titel": "",
            "OE-ID Signatur bei Routing 2.3": "",
            "OD-ID": "",
            "OD-ID Datenschutz-URL": "",
            "OD-ID URL": "",
        }

    # --------------------------------------------------------
    # Schritt 2:
    # OD-IDs
    # --------------------------------------------------------

    od_ids = find_od_ids(
        online_data
    )

    # --------------------------------------------------------
    # Schritt 3:
    # LBID
    # --------------------------------------------------------

    lbid = find_lbid(
        online_data,
        leika
    )

    # --------------------------------------------------------
    # Schritt 4:
    # Leistungsdetail
    # --------------------------------------------------------

    service_detail = None

    if lbid:

        service_detail = (
            get_service_detail(
                lbid,
                ars
            )
        )

    # Falls v6 keinen Treffer liefert,
    # verwenden wir zumindest die Daten
    # der Onlinedienst-Abfrage.

    source_for_oe = (
        service_detail
        if service_detail
        else online_data
    )

    # --------------------------------------------------------
    # Schritt 5:
    # OE
    # --------------------------------------------------------

    oe_data = collect_oe_data(
        ars,
        lbid,
        source_for_oe,
        od_ids
    )

    # --------------------------------------------------------
    # Schritt 6:
    # OD
    # --------------------------------------------------------

    od_data = collect_od_data(
        od_ids
    )

    return {
        "Leika-ID":
            leika,

        "OE-ID":
            "\n".join(
                oe_data["oe_ids"]
            ),

        "OE-ID Titel":
            "\n".join(
                oe_data["oe_titles"]
            ),

        "OE-ID Signatur Titel":
            "\n".join(
                oe_data[
                    "signature_titles"
                ]
            ),

        "OE-ID Signatur bei Routing 2.3":
            "\n".join(
                oe_data[
                    "routing_signatures"
                ]
            ),

        "OD-ID":
            "\n".join(
                od_data["od_ids"]
            ),

        "OD-ID Datenschutz-URL":
            "\n".join(
                od_data[
                    "privacy_urls"
                ]
            ),

        "OD-ID URL":
            "\n".join(
                od_data[
                    "od_urls"
                ]
            ),
    }


# ============================================================
# EXCEL
# ============================================================

def create_excel(
    results
):

    workbook = openpyxl.Workbook()

    worksheet = workbook.active

    worksheet.title = (
        "PVOG Resultate"
    )

    headers = [
        "Leika-ID",
        "OE-ID",
        "OE-ID Titel",
        "OE-ID Signatur Titel",
        "OE-ID Signatur bei Routing 2.3",
        "OD-ID",
        "OD-ID Datenschutz-URL",
        "OD-ID URL",
    ]

    worksheet.append(
        headers
    )

    # Header formatieren
    for cell in worksheet[1]:

        cell.font = Font(
            bold=True
        )

        cell.alignment = Alignment(
            vertical="top",
            wrap_text=True
        )

    # Daten
    for result in results:

        worksheet.append([
            result.get(
                "Leika-ID",
                ""
            ),

            result.get(
                "OE-ID",
                ""
            ),

            result.get(
                "OE-ID Titel",
                ""
            ),

            result.get(
                "OE-ID Signatur Titel",
                ""
            ),

            result.get(
                "OE-ID Signatur bei Routing 2.3",
                ""
            ),

            result.get(
                "OD-ID",
                ""
            ),

            result.get(
                "OD-ID Datenschutz-URL",
                ""
            ),

            result.get(
                "OD-ID URL",
                ""
            ),
        ])

    # Zeilenumbruch
    for row in worksheet.iter_rows():

        for cell in row:

            cell.alignment = Alignment(
                vertical="top",
                wrap_text=True
            )

    # Spaltenbreiten
    for column in worksheet.columns:

        maximum = 0

        column_number = (
            column[0].column
        )

        for cell in column:

            if cell.value is None:
                continue

            # Mehrzeilige Zellen berücksichtigen
            lines = str(
                cell.value
            ).split("\n")

            length = max(
                len(line)
                for line in lines
            )

            maximum = max(
                maximum,
                length
            )

        width = min(
            max(
                maximum + 2,
                12
            ),
            55
        )

        worksheet.column_dimensions[
            get_column_letter(
                column_number
            )
        ].width = width

    worksheet.freeze_panes = "A2"

    worksheet.auto_filter.ref = (
        worksheet.dimensions
    )

    output = BytesIO()

    workbook.save(
        output
    )

    output.seek(0)

    return output


# ============================================================
# STREAMLIT BENUTZEROBERFLÄCHE
# ============================================================

ars = st.text_input(
    "ARS-Nummer",
    max_chars=12,
    placeholder="12-stellige ARS eingeben",
)

start = st.button(
    "🔍 PVOG-Abfrage starten",
    type="primary",
    use_container_width=True,
)


if start:

    ars = ars.strip()

    # --------------------------------------------------------
    # Eingabeprüfung
    # --------------------------------------------------------

    if not ars:

        st.error(
            "Bitte eine ARS-Nummer eingeben."
        )

        st.stop()

    if not ars.isdigit():

        st.error(
            "Die ARS darf nur Ziffern enthalten."
        )

        st.stop()

    if len(ars) != 12:

        st.error(
            "Bitte eine 12-stellige ARS eingeben."
        )

        st.stop()

    # --------------------------------------------------------
    # Fortschritt
    # --------------------------------------------------------

    progress = st.progress(0)

    status = st.empty()

    results = []

    total = len(
        LEIKA_IDS
    )

    # --------------------------------------------------------
    # Alle sieben LeiKa
    # --------------------------------------------------------

    for index, leika in enumerate(
        LEIKA_IDS
    ):

        status.info(
            f"PVOG-Abfrage für LeiKa "
            f"{leika} "
            f"({index + 1}/{total})"
        )

        try:

            result = analyse_leika(
                ars,
                leika
            )

            results.append(
                result
            )

        except Exception as error:

            # Bei einem einzelnen Fehler
            # bleibt die Zeile erhalten.

            results.append({

                "Leika-ID":
                    leika,

                "OE-ID": "",

                "OE-ID Titel": "",

                "OE-ID Signatur Titel":
                    "",

                "OE-ID Signatur bei Routing 2.3":
                    "",

                "OD-ID": "",

                "OD-ID Datenschutz-URL":
                    "",

                "OD-ID URL": "",
            })

            st.warning(
                f"Fehler bei {leika}: "
                f"{error}"
            )

        progress.progress(
            int(
                ((index + 1) / total)
                * 100
            )
        )

    # --------------------------------------------------------
    # Excel
    # --------------------------------------------------------

    status.success(
        "PVOG-Abfrage abgeschlossen."
    )

    excel_file = create_excel(
        results
    )

    st.subheader(
        "Ergebnis"
    )

    # Vorschau
    st.dataframe(
        results,
        use_container_width=True,
        hide_index=True,
    )

    # Download
    st.download_button(
        label="📥 Excel-Datei herunterladen",
        data=excel_file,
        file_name=(
            f"PVOG_Resultate_{ars}.xlsx"
        ),
        mime=(
            "application/vnd.openxmlformats-officedocument."
            "spreadsheetml.sheet"
        ),
        use_container_width=True,
    )
